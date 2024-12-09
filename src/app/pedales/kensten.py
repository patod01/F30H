def kensten() -> (list, int):
     ("Returns a list of discarded projects and the number of \
     formatted projects")

     import pandas as pd, os, time, datetime
     from openpyxl import load_workbook
     from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
     from openpyxl.drawing.image import Image

     ### Setup constants ###
     path = 'AQUI'
     file = 'a.xlsx'

     anvil = f'{path}{os.sep}anvil'

     logo = Image(f'settings{os.sep}logo.png')

     ### Internal constants ###
     NARANJA = 'E4610F'
     NEGRO = '0C0808'
     BLANCO = 'FFFFFF'

     MESES = {
          1: 'Enero',
          2: 'Febrero',
          3: 'Marzo',
          4: 'Abril',
          5: 'Mayo',
          6: 'Junio',
          7: 'Julio',
          8: 'Agosto',
          9: 'Septiembre',
          10: 'Octubre',
          11: 'Noviembre',
          12: 'Diciembre',
     }

     def fmuer(E2: int) -> str:
          ('Convert e to p t converter and fomr.')
          E1: int = 34687
          P1: int = 787806000
          sges: int = 86400
          P2: int = (E2 - E1)*sges + P1 + 43200
          return time.strftime('%d/%m/%Y', time.localtime(P2))

     print('reading projects...')

     resumenes = pd.read_excel(file, sheet_name='resumenes')
     personal = pd.read_excel(file, sheet_name='personal')

     today = datetime.date.today()

     if not os.path.isdir(anvil):
          os.mkdir(anvil)

     skipped = list(personal[pd.isnull(personal.Nombre)].proyecto)
     print(
          'These projects are empty and will be skipped:\n', 
          skipped,
          sep=''
     )

     len(personal[pd.notnull(personal.Nombre)].proyecto.unique())
     print('|' + '-'*11 + ' progress ' + '-'*10 + '|')

     def progress(total, current, bar_length, old_bars):
          actual_progress = current/total
          bars = round(actual_progress*bar_length)
          if bars != old_bars:
               print('#', end='', flush=True)
          return bars

     current = 0
     bars = 0

     for project in personal[pd.notnull(personal.Nombre)].proyecto.unique():#[0:1]:
          ### Select project's data from dataframe ###
          reporte_a = resumenes[resumenes.proyecto == project][['atributo', 'dato']]
          reporte_b = personal[personal.proyecto == project][['Nombre', 'RUT']]

          reporte_a.iloc[3, 1] = fmuer(reporte_a.iloc[3, 1])
          reporte_a.iloc[4, 1] = fmuer(reporte_a.iloc[4, 1])

          ### Save unformatted xlsx ###
          report_name = f'F30-1 {project} {today.year - 1 if today.month == 1 else today.year} {12 if today.month == 1 else today.month - 1:#02}.xlsx'

          with pd.ExcelWriter(f'{anvil}{os.sep}{report_name}') as writer:
               reporte_a.to_excel(writer, sheet_name='Hoja1', index=False, header=False, startrow=2)
               reporte_b.to_excel(writer, sheet_name='Hoja1', index=False, startrow=len(reporte_a)+5)

          ### Init of xlsx formatting ###
          libro = load_workbook(f'{anvil}{os.sep}{report_name}')
          hoja = libro.active

          ### Ancho de columnas y primera fila ###
          (hoja.column_dimensions['A'].width) = 33
          (hoja.column_dimensions['B'].width) = 33

          hoja.add_image(logo, 'A1')

          hoja['b1'] = 'Horas informadas por recurso'
          hoja['b1'].font = Font(bold=True, color=NEGRO)

          ### Formato resumen proyecto ###
          for celda in hoja['A'][2:9]:
               celda.font = Font(bold=True)

          for celda in hoja['B'][2:9]:
               celda.alignment = Alignment(horizontal='left')

          hoja['A'][10].value = f'{MESES[12 if today.month == 1 else today.month - 1]} {today.year - 1 if today.month == 1 else today.year}'

          ### Formato tabla de recursos ###
          borde = Border(
               left=Side(style='thin', color=NEGRO),
               right=Side(style='thin', color=NEGRO),
               top=Side(style='thin', color=NEGRO),
               bottom=Side(style='thin', color=NEGRO)
          )

          for celda in hoja['A'][13:]:
               celda.border = borde

          for celda in hoja['B'][13:]:
               celda.border = borde
               celda.alignment = Alignment(horizontal='center')

          for celda in hoja['A'][12], hoja['B'][12]:
               celda.fill = PatternFill(start_color=NARANJA, end_color=NARANJA, fill_type='solid')
               celda.font = Font(bold=True, color=BLANCO)

          ### End ###
          hoja.sheet_view.showGridLines = False
          libro.save(f'{anvil}{os.sep}{report_name}')

          current+=1
          bars = progress(
               len(personal[pd.notnull(personal.Nombre)].proyecto.unique()),
               current,
               33,
               bars
          )

     print(' 100%')
     return skipped, current
